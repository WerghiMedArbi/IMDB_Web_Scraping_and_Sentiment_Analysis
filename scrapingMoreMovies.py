from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.workbook import Workbook
from lxml import etree
from time import sleep
import pandas as pd
import datetime

def getTime():
    return datetime.datetime.now().strftime("%H:%M:%S")

def scrapeReviews(url):
    df= pd.DataFrame(columns=["movie",'user',"rate","title","review","url"])
    # url = "https://www.imdb.com/title/tt9114286/reviews?spoiler=hide&sort=userRating&dir=desc&ratingFilter=0"
    path= ""
    while True:        
        source = requests.get(url)
        source.raise_for_status()
        soup = BeautifulSoup(source.text, 'html.parser')
        dom = etree.HTML(str(soup))
        if path =="": 
            path= dom.xpath('//div[@class="load-more-data"]/@data-ajaxurl')[0]
            movieName=dom.xpath('//h3[@itemprop="name"]/a/text()')[0]

        user = dom.xpath('//span[@class="display-name-link"]/a/text()')
        reviewRate = dom.xpath('//span[@class="point-scale"]/preceding-sibling::span/text()')
        reviewTitle = dom.xpath('//a[@class="title"]/text()') 
        reviewContent = [revC.text for revC in dom.xpath('//div[contains(@class,"text show-more__control")]')]
        reviewLink= [f"https://www.imdb.com{x}" for x in dom.xpath('//a[@class="title"]/@href')]

        for a in range(0,len(reviewLink)-len(reviewRate)):
            reviewRate.extend(["null"])
            
        try:
            paginationKey = dom.xpath('//div[@class="load-more-data"]/@data-key')[0]
        except IndexError:
            print(getTime()+':\t\x1b[4;32;41m Finished ' + movieName + '\x1b[0m')
            break
        url= f"https://www.imdb.com{path}&paginationKey={paginationKey}"  
        # print(movieName,url)
        dff = pd.DataFrame({
            "movie":movieName,
            'user':user,
            "rate":reviewRate,
            "title":reviewTitle,
            "review":reviewContent,
            "url":reviewLink
        })
        df= pd.concat([df,dff], ignore_index=True,sort=False)
    return df

# wb = Workbook()
# ws1 = wb.create_sheet("Sheet_A")
# ws1.title = "IMDB Movies"
# ws1.append(['rank', 'Name', 'Year', 'Rating', 'Metascore', 'Runtime', 'Genre', 'Votes'])
# ws2 = wb.create_sheet("Sheet_B", 0)
# ws2.title = "IMDB Reviews"
# ws2.append(["movie",'user',"rate","title","review","url"])

headers =False
pd.DataFrame(columns=['rank', 'Name', 'Year', 'Rating', 'Metascore', 'Runtime', 'Genre', 'Votes']).to_csv("movieDetails.csv",mode="a",index=False,header=headers)
scrapedMovieNames= pd.read_csv("movieDetails.csv")["Name"].tolist()

start = 1
# for start in range(1,1001):
while start <= 8900:
    url = f"https://www.imdb.com/search/title/?title_type=feature&user_rating=1.0,10.0&genres=sci-fi&count=250&start={start}&ref_=adv_nxt"
    print(getTime()+':\t\x1b[4;36;40m' + url + '\x1b[0m')

    start += 250
    source = requests.get(url)
    source.raise_for_status()
    soup = BeautifulSoup(source.text, 'html.parser')

    movies = soup.find('div', class_="lister-list").find_all('div', class_="lister-item mode-advanced")
    for movie in movies:
        name = movie.find('h3', class_="lister-item-header").a.text
        if (name in scrapedMovieNames):
            print(getTime()+':\t\x1b[6;35;40mAlready Scraped:\t\t\t' + name + '..\x1b[0m') # purple on black
            continue 
        print(getTime()+':\t\x1b[6;32;40m' + name + '\x1b[0m', name in scrapedMovieNames)# green on black

        movieId= movie.find('h3', class_="lister-item-header").a.get('href').split("/")[2]
        reviewLink=f"https://www.imdb.com/title/{movieId}/reviews?spoiler=hide&sort=userRating&dir=desc&ratingFilter=0"
        source = requests.get(reviewLink)
        source.raise_for_status()
        soup = BeautifulSoup(source.text, 'html.parser')
        
        scrapeReviews(reviewLink).to_csv("IMDB_reviews.csv", mode="a", index=False, header=headers)
        headers=False
        ## Movie details scraping 
        # ws2.append(scrapeReviews(reviewLink))

        try: 
            rank = movie.find('h3', class_="lister-item-header").get_text(strip=True).split('.')[0]
        except AttributeError:
            rank =""
            
        try:
            year = movie.find('span', class_="lister-item-year text-muted unbold").text[-5:-1] # famma mass houni idk why
        except AttributeError:
            year= ""
        try:
            genre = movie.find('span', class_="genre").get_text(strip=True)
        except AttributeError:
            genre= ""
        try:
            rating = movie.find('div', class_="ratings-bar").div.text
        except AttributeError:
            rating= ""
        try:
            runtime = movie.find('span', class_="runtime").get_text(strip=True)
        except AttributeError:
            runtime= ""
        try:
            votes = movie.find('span', {'name': 'nv'}).text
        except AttributeError:
            votes= ""

        # f metascore fama des films maandhomch metascore ---------------------------
        try:
            metascore = movie.find('div', class_="inline-block ratings-metascore").span.text
        except AttributeError:
            metascore = ""


        print(f'\x1b[4;37;40m {rank, name, year, rating, metascore, runtime, genre, votes} \x1b[0m')#undelined white on black
        # print({rank, name, year, rating, metascore, runtime, genre, votes})
        
        dict ={ 'rank':rank,
                'Name':name,
                'Year':year,
                'Rating':rating,
                'Metascore':metascore,
                'Runtime':runtime,
                'Genre':genre,
                'Votes':votes}
        pd.DataFrame([dict]).to_csv("movieDetails.csv",mode="a", index=False,header=False)












# except Exception as e:
#     print(getTime()+':\t\x1b[4;32;41m' + str(e) + '\x1b[0m')

# wb.save(filename='TooMany.xlsx')
# g4w6rcjqsr3ximsatx5dz2dezxzyyrrt3undz67napjzo6uynqwhlb674gzfnfu4ntdhkaa