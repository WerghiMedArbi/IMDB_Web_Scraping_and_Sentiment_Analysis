from bs4 import BeautifulSoup
import requests
import openpyxl
from openpyxl.workbook import Workbook


wb = Workbook()
 
ws1 = wb.create_sheet("Sheet_A")
ws1.title = "IMDB Movies"
 
ws2 = wb.create_sheet("Sheet_B", 0)
ws2.title = "IMDB Reviews"

try:
    
    page = 1
    while page != 1001:
        url = f"https://www.imdb.com/search/title/?groups=top_1000&sort=user_rating,asc&start={page}"
        print(url)
    
        page = page + 50
        source = requests.get(url)
    
        source.raise_for_status()

        soup = BeautifulSoup(source.text,'html.parser')
        
        

        movies = soup.find('div', class_= "lister-list") .find_all('div', class_="lister-item mode-advanced")
        
    
        for movie in movies:
            review_link = 'https://www.imdb.com'+movie.find('h3', class_= "lister-item-header").a.get('href')
            
            urlreview = review_link + "reviews?ref_=tt_ov_rt"
                    
            source = requests.get(urlreview)
            
            source.raise_for_status()
            
            soup = BeautifulSoup(source.text,'html.parser')
            
            reviews = soup.find('section', class_ = "article" ).find_all('div', class_ = "review-container")
            for review in reviews:
                
                user = review.find('span', class_="display-name-link").a.text
                try: 
                    rate = review.find('span', class_ ="rating-other-user-rating" ).text
                except AttributeError:
                    continue
                review_title = review.find('div',class_="lister-item-content").a.text
                try:
                    his_full_review = review.find('div', class_ = "text show-more__control").get_text(strip=True)
                except AttributeError:
                    review_title = None
                film_title = soup.find('h3', itemprop="name").a.get_text(strip=True)
                print(film_title)
                print(user,rate,review_title,his_full_review)
                ws2.append([film_title,user,rate,review_title,his_full_review])

            rank = movie.find('h3', class_= "lister-item-header").get_text(strip=True).split('.')[0]
            
            year = movie.find('span', class_= "lister-item-year text-muted unbold").text.strip('()')
            
            name = movie.find('h3', class_= "lister-item-header").a.text
            
            genre = movie.find('span', class_= "genre").get_text(strip=True) 
         
            rating = movie.find('div', class_= "ratings-bar").div.text 
       
            runtime = movie.find('span', class_= "runtime").get_text(strip=True)
            
            votes = movie.find('span',{'name':'nv'}).text
            
            #f metascore fama des films maandhomch metascore ---------------------------
            try:
                metascore = movie.find('div', class_= "inline-block ratings-metascore").span.text
            except AttributeError:
                metascore = None
            
            print(rank,name,year,rating,metascore,runtime,genre,votes) 
            
            ws1.append([rank,name,year,rating,metascore,runtime,genre,votes])
            '''
            
            sleep(randint(1,3))
            '''

       



except Exception as e:
    print(e)
   

    

wb.save(filename = 'IMDB_SCRAPING.xlsx')